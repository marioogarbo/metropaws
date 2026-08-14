import io
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload, selectinload

from database import get_db
import models, auth as auth_utils

router = APIRouter(prefix="/admin/exports", tags=["admin"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PESO_FORMAT = "#,##0.00"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
DATE_FORMAT = "yyyy-mm-dd"
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 45

# Philippine Standard Time has no daylight saving, so a fixed offset is safe
# and avoids requiring an IANA timezone database on the host.
PHILIPPINE_TIME = timezone(timedelta(hours=8))


@dataclass(frozen=True)
class ExportColumn:
    title: str
    number_format: Optional[str] = None


def _to_philippine_time(value: Optional[datetime]) -> Optional[datetime]:
    """openpyxl rejects timezone-aware datetimes, so convert then strip tzinfo."""
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(PHILIPPINE_TIME).replace(tzinfo=None)


def _centavos_to_pesos(centavos: Optional[int]) -> Optional[float]:
    if centavos is None:
        return None
    return centavos / 100


def _build_workbook(sheet_title: str, columns: list[ExportColumn], rows: list[list]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.freeze_panes = "A2"

    sheet.append([column.title for column in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        if column.number_format:
            for cell in sheet[letter][1:]:
                cell.number_format = column.number_format
        content_width = max(
            [len(column.title)] + [len(str(cell.value)) for cell in sheet[letter] if cell.value is not None]
        )
        sheet.column_dimensions[letter].width = min(max(content_width + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _xlsx_response(filename_prefix: str, buffer: io.BytesIO) -> StreamingResponse:
    stamp = datetime.now(PHILIPPINE_TIME).strftime("%Y-%m-%d")
    filename = f"{filename_prefix}_{stamp}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/members.xlsx")
def export_members(
    current_user: models.User = Depends(auth_utils.require_admin),
    db: Session = Depends(get_db),
):
    # DPA access log: exports contain members' personal data.
    print(f"[audit] admin={current_user.id} exported members to xlsx")
    members = (
        db.query(models.Member)
        .options(joinedload(models.Member.user), selectinload(models.Member.pets))
        .order_by(models.Member.joined_at.desc())
        .all()
    )
    columns = [
        ExportColumn("First Name"),
        ExportColumn("Last Name"),
        ExportColumn("Email"),
        ExportColumn("Phone"),
        ExportColumn("Address"),
        ExportColumn("Plan"),
        ExportColumn("Founding Member"),
        ExportColumn("Pets"),
        ExportColumn("Joined (PHT)", DATETIME_FORMAT),
    ]
    rows = [
        [
            member.first_name,
            member.last_name,
            member.user.email if member.user else None,
            member.phone,
            member.address,
            member.plan_type,
            "Yes" if member.is_founding else "No",
            ", ".join(pet.name for pet in member.pets),
            _to_philippine_time(member.joined_at),
        ]
        for member in members
    ]
    return _xlsx_response("metropaws_members", _build_workbook("Members", columns, rows))


@router.get("/reimbursements.xlsx")
def export_reimbursements(
    status: Optional[str] = Query(None),
    current_user: models.User = Depends(auth_utils.require_admin),
    db: Session = Depends(get_db),
):
    # DPA access log: exports contain members' financial data.
    print(f"[audit] admin={current_user.id} exported reimbursements to xlsx (status={status})")
    query = db.query(models.Reimbursement).options(
        joinedload(models.Reimbursement.service_type),
        joinedload(models.Reimbursement.pet),
        joinedload(models.Reimbursement.member).joinedload(models.Member.user),
    )
    if status:
        try:
            query = query.filter(models.Reimbursement.status == models.ReimbursementStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status value")
    claims = query.order_by(models.Reimbursement.created_at.desc()).all()
    columns = [
        ExportColumn("Member"),
        ExportColumn("Email"),
        ExportColumn("Pet"),
        ExportColumn("Service"),
        ExportColumn("Provider"),
        ExportColumn("Service Date", DATE_FORMAT),
        ExportColumn("Claimed (PHP)", PESO_FORMAT),
        ExportColumn("Approved (PHP)", PESO_FORMAT),
        ExportColumn("Status"),
        ExportColumn("Payout Method"),
        ExportColumn("Paid Reference"),
        ExportColumn("Admin Notes"),
        ExportColumn("Submitted (PHT)", DATETIME_FORMAT),
        ExportColumn("Reviewed (PHT)", DATETIME_FORMAT),
        ExportColumn("Paid (PHT)", DATETIME_FORMAT),
    ]
    rows = [
        [
            f"{claim.member.first_name} {claim.member.last_name}" if claim.member else None,
            claim.member.user.email if claim.member and claim.member.user else None,
            claim.pet.name if claim.pet else None,
            claim.service_type.name if claim.service_type else None,
            claim.provider_name,
            claim.service_date,
            _centavos_to_pesos(claim.claimed_amount_centavos),
            _centavos_to_pesos(claim.approved_amount_centavos),
            claim.status.value,
            claim.member.payout_method if claim.member else None,
            claim.paid_reference,
            claim.admin_notes,
            _to_philippine_time(claim.created_at),
            _to_philippine_time(claim.reviewed_at),
            _to_philippine_time(claim.paid_at),
        ]
        for claim in claims
    ]
    return _xlsx_response("metropaws_reimbursements", _build_workbook("Reimbursements", columns, rows))


@router.get("/founding-reservations.xlsx")
def export_founding_reservations(
    current_user: models.User = Depends(auth_utils.require_admin),
    db: Session = Depends(get_db),
):
    print(f"[audit] admin={current_user.id} exported founding reservations to xlsx")
    reservations = (
        db.query(models.FoundingReservation)
        .order_by(models.FoundingReservation.created_at.desc())
        .all()
    )
    columns = [
        ExportColumn("First Name"),
        ExportColumn("Last Name"),
        ExportColumn("Email"),
        ExportColumn("Phone"),
        ExportColumn("Barangay"),
        ExportColumn("Message"),
        ExportColumn("Status"),
        ExportColumn("Admin Notes"),
        ExportColumn("Submitted (PHT)", DATETIME_FORMAT),
    ]
    rows = [
        [
            reservation.first_name,
            reservation.last_name,
            reservation.email,
            reservation.phone,
            reservation.barangay,
            reservation.message,
            reservation.status.value,
            reservation.admin_notes,
            _to_philippine_time(reservation.created_at),
        ]
        for reservation in reservations
    ]
    return _xlsx_response(
        "metropaws_founding_reservations",
        _build_workbook("Founding Reservations", columns, rows),
    )
